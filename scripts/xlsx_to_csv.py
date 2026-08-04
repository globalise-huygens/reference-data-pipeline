#!/usr/bin/env python3
"""
Split an XLSX workbook into one CSV per sheet.

Usage:
    python3 xlsx_to_csv.py <input.xlsx> [output_dir]

The output directory defaults to "csv" next to the workbook.
"""

import argparse
import csv
import os
import warnings

from openpyxl import load_workbook

from utils import cell_to_text, safe_filename as sanitize_filename

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
)


def parse_args(args_list: list[str] | None = None) -> argparse.Namespace:
    """
    Parse CLI arguments for XLSX to CSV conversion.

    Args:
        args_list (list[str], optional): CLI arguments list for testing. Defaults to None.

    Returns:
        argparse.Namespace: Parsed command-line arguments.

    Examples:
        >>> parse_args(["book.xlsx"]).input_xlsx
        'book.xlsx'
    """
    parser = argparse.ArgumentParser(
        description="Split an XLSX workbook into one CSV per sheet."
    )
    parser.add_argument("input_xlsx")
    parser.add_argument("output_dir", nargs="?")
    return parser.parse_args(args_list)


def write_sheet_csv(sheet: object, output_path: str) -> int:
    """
    Write an openpyxl worksheet out to a CSV file.

    Args:
        sheet (object): Openpyxl worksheet instance.
        output_path (str): Filepath for target CSV file.

    Returns:
        int: Number of data rows written.
    """
    rows = list(sheet.iter_rows(values_only=True))  # type: ignore[attr-defined]
    if not rows:
        with open(output_path, "w", encoding="utf-8") as handle:
            handle.write("")
        return 0

    headers = list(rows[0])
    width = (
        max(len(headers), *(len(row) for row in rows[1:])) if rows[1:] else len(headers)
    )
    normalized_headers = [
        (
            cell_to_text(headers[idx])
            if idx < len(headers) and headers[idx] not in (None, "")
            else f"column_{idx + 1}"
        )
        for idx in range(width)
    ]

    written_rows = 0
    with open(output_path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(normalized_headers)

        for row in rows[1:]:
            padded = list(row[:width]) + [""] * max(0, width - len(row))
            if all(value in (None, "") for value in padded):
                continue
            writer.writerow([cell_to_text(value) for value in padded])
            written_rows += 1

    return written_rows


def main() -> int:
    """
    Execute main entry point for XLSX workbook splitting.

    Returns:
        int: Exit status code (0 for success, 1 for error).
    """
    args = parse_args()
    input_path = os.path.abspath(os.path.expanduser(args.input_xlsx))
    if not os.path.exists(input_path):
        print(f"Error: file '{input_path}' not found.")
        return 1

    output_dir = (
        os.path.abspath(os.path.expanduser(args.output_dir))
        if args.output_dir
        else os.path.join(os.path.dirname(input_path), "csv")
    )
    os.makedirs(output_dir, exist_ok=True)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    total_rows = 0
    written_files = []

    for sheet in workbook.worksheets:
        output_path = os.path.join(output_dir, f"{sanitize_filename(sheet.title)}.csv")
        rows_written = write_sheet_csv(sheet, output_path)
        total_rows += rows_written
        written_files.append(output_path)

    print(
        f"Wrote {len(written_files)} CSV files to '{output_dir}' ({total_rows} data rows)."
    )
    return 0


if __name__ == "__main__":
    import doctest

    doctest.testmod()
    raise SystemExit(main())
